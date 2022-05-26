# -*- coding: utf-8 -*-

import os
import pandas as pd
from openpyxl import load_workbook


def tabit(
    data,
    con_vars=None,
    cat_vars=None,
    clu_col="target",
    outfile=None,
    startcol=0,
    incl_total=True,
):
    """
    incl_total: ignore index and total col
    """
    mean = pd.DataFrame()
    if con_vars is not None:
        for i in con_vars:
            try:
                clu_col_mean = data[i].groupby(data[clu_col]).mean()
                total_mean = data[i].mean()
                clu_col_mean = pd.DataFrame(clu_col_mean)
                clu_col_mean = clu_col_mean.T
                clu_col_mean.insert(0, "total", total_mean)
                mean = mean.append(clu_col_mean)
            except Exception as e:
                # raise(e, ",check your data in column {}".format(i))
                raise TypeError(e, "check your data in column {}".format(i))

        mean.insert(0, "convar", mean.index)
        total_count = data[clu_col].count()
        clu_col_count = data[clu_col].groupby(data[clu_col]).count()
        base = pd.DataFrame(clu_col_count).T
        base.insert(0, "total", total_count)
        base.insert(0, "convar", "base")
        mean = pd.concat([base, mean], axis=0)
        name_a2 = data[clu_col].unique()
        name_a2.sort()
        x = 2
        for i in name_a2:
            name2 = mean.columns[x]
            x = x + 1
            stri = str(i)
            clu = "clu"
            mean.rename(columns={name2: clu + " " + stri}, inplace=True)
    if incl_total:
        mean = mean
    else:
        mean = mean.iloc[:, 2:]
    append_df_to_excel(
        outfile,
        mean,
        startrow=0,
        startcol=startcol,
        sheet_name="tables",
        index=False,
    )
    cat_startrow = mean.shape[0] + 2

    if cat_vars is not None:
        list1 = []
        new = pd.DataFrame()
        for i in cat_vars:
            total_percentage = (data[clu_col].groupby([data[i]]).count()) / data[
                i
            ].count()
            clu_col_percentage = (
                data[clu_col].groupby([data[i], data[clu_col]]).count()
            ) / data[clu_col].groupby([data[clu_col]]).count()
            clu_col_percentage = pd.DataFrame(clu_col_percentage)
            clu_col_percentage = clu_col_percentage.unstack(clu_col)
            cat = pd.concat([total_percentage, clu_col_percentage], axis=1)
            cat.fillna(0, inplace=True)
            name1 = cat.columns[0]
            cat.rename(columns={name1: "percentage"}, inplace=True)
            cat["name"] = cat.index
            cat["name"] = cat["name"].astype("str")
            name_a1 = i + " " + "=" + " " + cat["name"]
            cat.drop(columns=["name"], inplace=True)
            cat.insert(0, "catvars", name_a1)
            name_a2 = data[clu_col].unique()
            name_a2.sort()
            x = 2
            for i in name_a2:
                name2 = cat.columns[x]
                x = x + 1
                stri = str(i)
                clu = "clu"
                cat.rename(columns={name2: clu + " " + stri}, inplace=True)
            new = pd.concat([new, cat])
            list1.append(cat)
        total_count = data[clu_col].count()
        clu_col_count = data[clu_col].groupby(data[clu_col]).count()
        base = pd.DataFrame(clu_col_count).T
        base.insert(0, "percentage", total_count)
        base.insert(0, "catvars", "base")
        name_a2 = data[clu_col].unique()
        name_a2.sort()
        x = 2
        for i in name_a2:
            name2 = base.columns[x]
            x = x + 1
            stri = str(i)
            clu = "clu"
            base.rename(columns={name2: clu + " " + stri}, inplace=True)
        new = pd.concat([base, new], axis=0)
        cat_startrow = cat_startrow if con_vars is not None else 0
        if incl_total:
            new = new
        else:
            new = new.iloc[:, 2:]
        append_df_to_excel(
            outfile,
            new,
            startrow=cat_startrow,
            startcol=startcol,
            sheet_name="tables",
            index=False,
        )
        return new.shape[1] + 1


def append_df_to_excel(
    filename,
    df,
    sheet_name="Sheet1",
    startrow=None,
    startcol=None,
    truncate_sheet=False,
    **to_excel_kwargs
):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs
        )
        return

    # ignore [engine] parameter if it was passed
    if "engine" in to_excel_kwargs:
        to_excel_kwargs.pop("engine")

    writer = pd.ExcelWriter(
        filename, engine="openpyxl", mode="a", if_sheet_exists="overlay"
    )

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # # get the last row in the existing Excel sheet
    # # if it was not specified explicitly
    # if startrow is None and sheet_name in writer.book.sheetnames:
    #     startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(
        writer,
        sheet_name=sheet_name,
        startrow=startrow,
        startcol=startcol,
        **to_excel_kwargs
    )

    # save the workbook
    writer.save()


if __name__ == "__main__":
    data = pd.read_excel("../tests/midea_test/data/midea_data_for_segment.xlsx")

    start_col = 0
    for i in range(3):
        start_col += tabit(
            data=data,
            con_vars=[
                "S9Money",
                "B2_1",
                "B2_2",
                "B2_3",
                "B2_4",
                "B2_5",
                "B2_6",
                "B2_7",
                "B2_8",
                "B2_9",
                "B2_10",
                "B2_11",
                "B2_12",
                "B2_13",
                "O1_Loop_1_O1",
                "O1_Loop_2_O1",
                "O1_Loop_3_O1",
                "O1_Loop_4_O1",
                "O1_Loop_5_O1",
                "O1_Loop_6_O1",
                "O1_Loop_7_O1",
                "O1_Loop_8_O1",
                "O1_Loop_9_O1",
                "O1_Loop_10_O1",
                "O1_Loop_11_O1",
                "O1_Loop_12_O1",
                "O1_Loop_13_O1",
                "O1_Loop_14_O1",
                "O1_Loop_15_O1",
                "O1_Loop_16_O1",
                "C1_Loop_1_C1",
                "C1_Loop_2_C1",
                "C1_Loop_3_C1",
                "C1_Loop_4_C1",
                "C1_Loop_5_C1",
                "C1_Loop_6_C1",
                "C1_Loop_7_C1",
                "C1_Loop_8_C1",
                "C1_Loop_9_C1",
                "C1_Loop_10_C1",
                "C1_Loop_11_C1",
                "C1_Loop_12_C1",
            ],
            cat_vars=[
                "lifestage",
                "lifestage3",
                "income_2",
                "E2a",
                "E2b",
                "E2c",
                "E2d",
                "N1_Loop_1_N1",
                "N1_Loop_2_N1",
                "N1_Loop_3_N1",
                "N1_Loop_4_N1",
                "N1_Loop_5_N1",
                "N1_Loop_6_N1",
                "C11",
                "T8",
                "T9",
                "B1_Loop_1_B1",
            ],
            clu_col="final_target",
            outfile="./tabit.xlsx",
            startcol=start_col,
        )
